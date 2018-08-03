# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime

class PeoplespiderPipeline(object):

    def __init__(self):

        today = datetime.datetime.now()
        self.wb = load_workbook('D:/news_data/%s.xlsx' % (today.strftime('%m%d')))
        # self.wb = Workbook()
        self.ws = self.wb.active
        # self.ws.append(['新闻标题', '发布时间', '新闻来源', '新闻作者', '新闻内容', '留言评论'])
        # today = datetime.datetime.now()
        # self.wb = load_workbook('D:/news_data/%s.xlsx' % (today.strftime('%m%d')))
        # self.ws = self.wb.active

    def process_item(self, item,spider):
        line = [item['news_title'], item['news_time'], item['news_source'], item['news_author'],
                item['news_content'], item['news_comment'], ]

        self.ws.append(line)
        return item

    def close_spider(self,spider):
        today = datetime.datetime.now()
        self.wb.save('D:/news_data/%s.xlsx' % (today.strftime('%m%d')))
